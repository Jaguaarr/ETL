"""Bank Al Maghrib scraper.

Ce script télécharge des rapports PDF (et, pour certaines sections, des
fichiers Excel) publiés sur des pages de statistiques sélectionnées de
Bank Al-Maghrib et en extrait les lignes de tableau vers CSV/JSON/SQLite.

Usage (une section) :
    python scraper_bkam.py --section regional_credit --output ../../../datasets/bkm/bkam_credit_regional.csv --normalize

Usage (toutes les sections retenues, cf. ALL_SECTIONS) :
    python scraper_bkam.py --all
    python scraper_bkam.py --all --dry-run

Sections supportées :
    regional_credit                  Répartition régionale (rayons d'action) des guichets/dépôts/crédits
    dashboard_credits_depots         Tableau de bord crédits-dépôts bancaires (national) — EXCLU de --all :
                                      pas de grain temporel/géo exploitable pour ce pipeline.
    credits_depots_localites         Répartition par localités (villes) des guichets/dépôts/crédits
    credit_objet_economique          Crédit bancaire par objet éco. (immobilier, équipement,
                                      trésorerie, consommation) — série statistique monétaire n°12
    credit_secteur_institutionnel    Crédit bancaire par secteur institutionnel (ménages,
                                      sociétés non financières privées/publiques) — série n°13
    densite_bancaire                 Nombre d'agences bancaires + densité bancaire, extraits du
                                      texte (pas d'un tableau) du dernier Rapport annuel de supervision
                                      bancaire

--all boucle sur ALL_SECTIONS, écrit un CSV par section (nom de fichier fixe,
cf. SECTION_OUTPUT_FILENAMES) dans datasets/bkm/, toujours avec --normalize
(le SQL bronze correspondant attend des en-têtes snake_case) — c'est ce que
scripts/bkm/pipeline.py appelle.
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import re
import sqlite3
import time
import unicodedata
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

import pdfplumber
import requests
from bs4 import BeautifulSoup
from requests import RequestException, Session

BASE_URL = "https://www.bkam.ma"
DEFAULT_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
}

# Retry policy for flaky network calls to bkam.ma
MAX_RETRIES = 3
RETRY_BACKOFF_SECONDS = 2.0

logger = logging.getLogger("bank_almaghreb_scraper")

SECTION_PAGES = {
    "regional_credit": {
        "label": "Regional deposits and credits by agency action radius",
        "url": (
            "https://www.bkam.ma/Statistiques/Statistiques-sur-le-secteur-bancaire/"
            "Repartition-regionale/Repartition-des-guichets-des-depots-et-des-credits-"
            "par-decaissement-des-banques-par-rayons-d-action-des-agences-de-bam"
        ),
        "filter_keywords": [
            "repartition", "depots", "credits", "rayon", "region",
        ],
        "kind": "pdf_table",
    },
    "dashboard_credits_depots": {
        "label": "Credits and deposits dashboard",
        "url": (
            "https://www.bkam.ma/Statistiques/Chiffres-cles-de-l-economie-nationale/"
            "Credits-et-depots-bancaires/Tableau-de-bord-credits-depots-bancaires-2026"
        ),
        "filter_keywords": ["flash", "credits", "depots"],
        "kind": "pdf_table",
    },
    # -------------------------------------------------------------------
    # NOUVEAU : répartition par localités (villes), page sœur de
    # regional_credit -- même structure de tableau (Code / Localité /
    # Guichets / Dépôts Montant+% / Crédits Montant+%), donc réutilise le
    # même pipeline d'extraction PDF sans modification.
    # -------------------------------------------------------------------
    "credits_depots_localites": {
        "label": "Répartition par localités des guichets, dépôts et crédits",
        "url": (
            "https://www.bkam.ma/Statistiques/Statistiques-sur-le-secteur-bancaire/"
            "Repartition-regionale/Ventilation-par-localites-des-guichets-des-depots-"
            "et-des-credits-par-decaissement-des-banques"
        ),
        "filter_keywords": ["repartition", "localites", "depots", "credits"],
        "kind": "pdf_table",
    },
    # -------------------------------------------------------------------
    # NOUVEAU : crédit bancaire par objet économique (immobilier,
    # équipement, trésorerie, consommation) -- série statistique
    # monétaire n°12. Page mère : Series-statistiques-monetaires.
    # Le format de fichier (pdf vs xlsx) n'a pas pu être confirmé
    # sans accès réseau à bkam.ma -- fetch_pdf_links ne remonte que les
    # liens .pdf par défaut ; ajuster ACCEPTED_EXTENSIONS ci-dessous si le
    # fichier réel est un .xlsx.
    # -------------------------------------------------------------------
    "credit_objet_economique": {
        "label": "Ventilation du crédit bancaire par objet économique",
        "url": "https://www.bkam.ma/Statistiques/Statistiques-monetaires/Series-statistiques-monetaires",
        # Phrase complete (pas juste "objet economique") : cette page liste
        # ~40 series statistiques monetaires, dont l'item 39 "Creances des
        # ASF sur les agents non financiers PAR OBJET ECONOMIQUE" qui
        # contient aussi "objet economique" mais n'a rien a voir (verifie en
        # direct : "objet economique" seul matchait a tort les 2 items).
        "filter_keywords": ["credit bancaire par objet economique"],
        # Exclut l'item 14 "Ventilation croisee du credit bancaire par objet
        # economique ET par secteur institutionnel" (meme page, croise 2
        # dimensions -- pas le meme schema de colonnes que l'item 12 dedie).
        "exclude_keywords": ["croisee"],
        "kind": "pdf_or_xlsx_table",
    },
    "credit_secteur_institutionnel": {
        "label": "Ventilation du crédit bancaire par secteur institutionnel",
        "url": "https://www.bkam.ma/Statistiques/Statistiques-monetaires/Series-statistiques-monetaires",
        "filter_keywords": ["secteur institutionnel"],
        "exclude_keywords": ["croisee"],  # cf. commentaire credit_objet_economique ci-dessus
        "kind": "pdf_or_xlsx_table",
    },
    # -------------------------------------------------------------------
    # NOUVEAU : densité bancaire + nombre d'agences -- PAS un tableau,
    # c'est une phrase dans le texte narratif du Rapport annuel de
    # supervision bancaire (page listant les rapports par année). On va
    # chercher le PDF le plus récent puis en extraire le texte (pas les
    # tableaux) pour y appliquer des regex.
    # -------------------------------------------------------------------
    "densite_bancaire": {
        "label": "Densité bancaire et nombre d'agences (extrait du Rapport annuel de supervision bancaire)",
        # NB: "/Supervision-bancaire/Publications" (URL initiale) est une
        # page de navigation SANS lien PDF direct -- verifie en direct
        # (0 lien PDF/XLSX trouve). Les rapports reels vivent sous
        # Publications-et-recherche/Publications-institutionnelles/...
        "url": "https://www.bkam.ma/Publications-et-recherche/Publications-institutionnelles/Rapport-annuel-sur-la-supervision-bancaire",
        # Le texte des liens est generique ("(PDF)") -- le filtre matche sur
        # le nom de fichier ("Rapport DSB 2024.pdf", etc.), pas sur le texte
        # du lien.
        "filter_keywords": ["rapport"],
        "kind": "pdf_text_regex",
    },
}

# Sections scrapées par --all (donc par scripts/bkm/pipeline.py). Exclut
# "dashboard_credits_depots" : page de tableau de bord national sans grain
# temporel/géographique exploitable par le modèle bronze/silver/gold de ce
# pipeline (documenté ici plutôt que scrapé "au cas où").
ALL_SECTIONS = [
    "regional_credit",
    "credits_depots_localites",
    "credit_objet_economique",
    "credit_secteur_institutionnel",
    "densite_bancaire",
]

# Nom de fichier de sortie fixe par section, attendu par le SQL bronze
# correspondant (scripts/bkm/sql/bronze/05-09_ddl_load_*.sql) — un fichier
# par section, jamais le fichier générique bank_almaghreb_data.csv.
SECTION_OUTPUT_FILENAMES = {
    "regional_credit": "bkam_credit_regional.csv",
    "credits_depots_localites": "bkam_credit_localites.csv",
    "credit_objet_economique": "bkam_credit_objet_eco.csv",
    "credit_secteur_institutionnel": "bkam_credit_secteur.csv",
    "densite_bancaire": "bkam_densite_bancaire.csv",
}


@dataclass
class PdfReportLink:
    title: str
    url: str


@dataclass
class ScrapedRow:
    report_title: str
    report_url: str
    pdf_filename: str
    page_number: int
    row_number: int
    data: dict[str, Any]


def normalize_text(text: str) -> str:
    text = text or ""
    text = unicodedata.normalize("NFKC", text)
    text = text.strip()
    return text


def normalize_header_cell(cell: str) -> str:
    return normalize_text(cell).replace("\n", " ").strip()


def normalize_number(value: str) -> Any:
    if value is None:
        return None
    value = normalize_text(value)
    if value == "":
        return None

    is_percent = value.endswith("%")
    value = value.replace("%", "")
    value = value.replace("\u00a0", " ")
    value = value.replace(" ", "")
    value = value.replace("\u202f", "")
    value = value.replace("\u2009", "")
    value = value.replace("\u200b", "")
    value = value.replace('"', "")
    value = value.replace("'", "")
    value = value.replace(",", ".")

    if value == "":
        return None

    if is_percent:
        try:
            return float(value)
        except ValueError:
            return value

    if re.fullmatch(r"[-+]?[0-9]+(\.[0-9]+)?", value):
        if "." in value:
            return float(value)
        return int(value)

    return value


def normalize_string(value: str) -> str:
    if value is None:
        return ""
    return normalize_text(value)


def remove_accents(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def normalize_field_name(label: str) -> str:
    label = normalize_text(label).lower()
    label = remove_accents(label)
    label = label.replace("%", " percent ")
    label = re.sub(r"[^a-z0-9]+", "_", label)
    label = re.sub(r"_+", "_", label).strip("_")
    if not label:
        return "column"

    normalized = label
    pattern_map = [
        (r"\bcode\b.*\brayon\b", "code_rayon_action"),
        (r"\brayon\b.*\baction\b", "rayon_action"),
        (r"\brayon\b", "rayon_action"),
        (r"\bcode\b.*\blocalite\b", "code_localite"),
        (r"\blocalite\b|\bville\b|\bcommune\b", "localite"),
        (r"\bnombre\b.*\bguichet\b", "nombre_guichets"),
        (r"\bguichet\b", "nombre_guichets"),
        (r"\bdepots?\b.*\bmontant\b", "depots_montant"),
        (r"\bdepots?\b.*\bpercent\b", "depots_percent"),
        (r"\bcredit\b.*\bmontant\b", "credits_montant"),
        (r"\bcredit\b.*\bpercent\b", "credits_percent"),
        # -- Nouveaux libellés pour credit_objet_economique --
        (r"\bimmobilier\b", "credit_immobilier"),
        (r"\bequipement\b", "credit_equipement"),
        (r"\btresorerie\b", "credit_tresorerie"),
        (r"\bconsommation\b", "credit_consommation"),
        # -- Nouveaux libellés pour credit_secteur_institutionnel --
        (r"\bmenages?\b", "credit_menages"),
        (r"\bsocietes?\b.*\bnon\b.*\bfinancieres?\b.*\bprivees?\b", "credit_entreprises_privees"),
        (r"\bsocietes?\b.*\bnon\b.*\bfinancieres?\b.*\bpubliques?\b", "credit_entreprises_publiques"),
        (r"\bentreprises?\b", "credit_entreprises"),
        (r"\bpercent\b", "percent"),
        (r"\bvariation\b.*\bpercent\b", "variation_percent"),
        (r"\bencours\b.*\bvariation\b", "variation_percent"),
    ]
    for pattern, canonical in pattern_map:
        if re.search(pattern, normalized):
            return canonical
    return normalized


def normalize_data_for_db(data: dict[str, Any]) -> dict[str, Any]:
    normalized_data: dict[str, Any] = {}
    for key, value in data.items():
        normalized_key = normalize_field_name(key)
        if isinstance(value, str):
            normalized_data[normalized_key] = normalize_number(value)
        else:
            normalized_data[normalized_key] = value
    return normalized_data


def extract_periode_from_title(title: str) -> str | None:
    """Extrait une période 'MM-AAAA' du titre du rapport (ex: 'Répartition
    par rayon d'action des dépôts et crédits 12-2025' -> '12-2025').
    Utilisé pour dater les lignes des rapports mensuels régionaux, dont le
    tableau lui-même ne contient pas de colonne date."""
    match = re.search(r"(0[1-9]|1[0-2])[-/](\d{4})", title)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    match = re.search(r"\b(\d{4})\b", title)
    if match:
        return match.group(1)
    return None


def build_csv_headers(rows: list[ScrapedRow]) -> list[str]:
    field_names: set[str] = set()
    for row in rows:
        field_names.update(row.data.keys())
    base_fields = ["report_title", "report_url", "pdf_filename", "page_number", "row_number"]
    return base_fields + sorted(field_names)


def retry_request(session: Session, method: str, url: str, **kwargs: Any) -> requests.Response:
    """GET/HEAD/etc with exponential backoff. Raises the last exception after MAX_RETRIES."""
    last_exc: Exception | None = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = session.request(method, url, **kwargs)
            response.raise_for_status()
            return response
        except RequestException as exc:
            last_exc = exc
            if attempt < MAX_RETRIES:
                wait = RETRY_BACKOFF_SECONDS * (2 ** (attempt - 1))
                logger.warning("Request to %s failed (attempt %d/%d): %s. Retrying in %.1fs",
                               url, attempt, MAX_RETRIES, exc, wait)
                time.sleep(wait)
    assert last_exc is not None
    raise last_exc


# Extensions de fichier téléchargeable reconnues pour la découverte de liens.
# "pdf_or_xlsx_table" a besoin des deux ; les autres sections ne veulent que
# des PDF (comportement d'origine, inchangé).
ACCEPTED_EXTENSIONS_BY_KIND = {
    "pdf_table": (".pdf",),
    "pdf_text_regex": (".pdf",),
    "pdf_or_xlsx_table": (".pdf", ".xlsx", ".xls"),
}


def fetch_pdf_links(
    url: str,
    session: Session,
    filter_keywords: list[str] | None = None,
    accepted_extensions: tuple[str, ...] = (".pdf",),
    exclude_keywords: list[str] | None = None,
) -> list[PdfReportLink]:
    response = retry_request(session, "GET", url, timeout=30)
    soup = BeautifulSoup(response.text, "html.parser")
    pdf_links: list[PdfReportLink] = []

    def has_any(text: str, keywords: list[str]) -> bool:
        # Accent-insensitive : les rapports recents sur bkam.ma utilisent des
        # noms de fichiers accentues ("Repartition par localites...") alors
        # que les archives plus anciennes utilisent des noms tout en
        # majuscules sans accents ("REPARTITION PAR LOCALITES..."). Une
        # comparaison ASCII stricte ne matche que les anciens (verifie en
        # direct contre bkam.ma : les rapports 2024-2026 etaient sinon
        # systematiquement ignores).
        normalized = remove_accents(normalize_text(text).lower())
        return any(remove_accents(normalize_text(kw).lower()) in normalized for kw in keywords)

    def matches_keyword(text: str) -> bool:
        return True if not filter_keywords else has_any(text, filter_keywords)

    def is_excluded(text: str) -> bool:
        return bool(exclude_keywords) and has_any(text, exclude_keywords)

    for anchor in soup.select("a[href]"):
        href = anchor["href"].strip()
        if not href.lower().endswith(accepted_extensions):
            continue
        title = normalize_text(anchor.get_text(separator=" "))
        full_url = href if href.startswith("http") else BASE_URL + href
        if is_excluded(title) or is_excluded(full_url):
            continue
        if matches_keyword(title) or matches_keyword(full_url):
            pdf_links.append(PdfReportLink(title=title or Path(full_url).name, url=full_url))

    unique_links: list[PdfReportLink] = []
    seen: set[str] = set()
    for link in pdf_links:
        if link.url in seen:
            continue
        seen.add(link.url)
        unique_links.append(link)

    return unique_links


def download_pdf(link: PdfReportLink, download_dir: Path, session: Session, verbose: bool = False) -> Path:
    download_dir.mkdir(parents=True, exist_ok=True)
    filename = normalize_text(Path(link.url).name)
    target = download_dir / filename
    if target.exists() and target.stat().st_size > 0:
        if verbose:
            logger.info("Using cached file: %s", target)
        return target

    if verbose:
        logger.info("Downloading %s", link.url)
    response = retry_request(session, "GET", link.url, timeout=60)

    is_pdf = response.content.startswith(b"%PDF")
    is_xlsx = response.content[:2] == b"PK"  # xlsx is a zip archive
    if not (is_pdf or is_xlsx):
        raise ValueError(
            f"Response for {link.url} does not look like a PDF or XLSX "
            f"(got content-type={response.headers.get('Content-Type')!r}); "
            "the link may be stale or bkam.ma returned an error page."
        )

    target.write_bytes(response.content)
    return target


def _is_subheader_row(row: list[str]) -> bool:
    """True if `row` looks like the 2nd half of a split header (e.g. the
    "Montant" / "%" sub-labels under a "Depots" / "Credits" parent header),
    as opposed to a real data row."""
    return any(
        cell and ("montant" in cell.lower() or "%" in cell or "dépôts" in cell.lower())
        for cell in row
    )


def merge_header_rows(rows: list[list[str]]) -> list[str] | None:
    """Merge a 1- or 2-row PDF table header into a single flat header row.

    BAM tables commonly split headers across two lines, e.g.:
        Row 1: ["Code", "Rayon d'action", "Depots", "Depots", "Credits", "Credits"]
        Row 2: ["",     "",               "Montant", "%",     "Montant", "%"]
    Empty cells in row 1 inherit the last non-empty "parent" label so that
    row 2 sub-headers (Montant / %) get merged into e.g. "Depots Montant".
    Other BAM tables (verifie en direct : rapport "localites") ont un
    en-tete a UNE seule ligne -- cf. header_row_count() qui determine lequel
    des deux cas s'applique avant d'appeler cette fonction.
    """
    if not rows:
        return None
    processed = [[normalize_header_cell(cell) for cell in row] for row in rows]
    if len(processed) == 1:
        return processed[0]

    first, second = processed[0], processed[1]

    if _is_subheader_row(second):
        max_len = max(len(first), len(second))
        merged = []
        last_parent = ""
        for i in range(max_len):
            part1 = first[i] if i < len(first) else ""
            part2 = second[i] if i < len(second) else ""
            if part1:
                last_parent = part1
            else:
                part1 = last_parent
            combined = " ".join([part1, part2]).strip()
            merged.append(combined or part1)
        return merged

    return first


def header_row_count(table: list[list[str]], header_index: int) -> int:
    """Combien de lignes a partir de header_index forment l'en-tete : 2 si
    la ligne suivante est un sous-en-tete (Montant/%/Depots), sinon 1.
    Determine dynamiquement (pas suppose fixe a 2) : un rapport BAM sur deux
    verifie en direct a un en-tete a une seule ligne, et supposer 2 a tort
    mangeait silencieusement le premier enregistrement de chaque page
    (consomme comme "2e ligne d'en-tete" au lieu d'etre une donnee)."""
    next_row = table[header_index + 1] if header_index + 1 < len(table) else None
    if next_row is not None and _is_subheader_row(next_row):
        return 2
    return 1


def find_header_index(rows: list[list[str]]) -> int | None:
    """Localise la ligne d'en-tete parmi les 3 premieres. Exige au moins 2
    mots-cles ET la presence de "depots"/"credit" (pas juste 1 mot-cle
    quelconque) pour eviter de prendre a tort pour un en-tete soit une ligne
    de donnees ordinaire, soit un tableau annexe sans rapport (verifie en
    direct sur le rapport "localites" : certaines pages contiennent, apres
    le tableau principal, un tableau complementaire "Localite / Nombre de
    guichets" sans montants -- hors perimetre de ce dataset -- et d'autres
    pages n'ont carrement pas de ligne d'en-tete du tout, juste des
    donnees). Retourne None si aucune ligne ne correspond : le tableau est
    alors ignore plutot que mal interprete."""
    keywords = ("rayon", "localite", "code", "d\u00e9p\u00f4ts", "cr\u00e9dit", "nombre")
    for index, row in enumerate(rows[:3]):
        row_text = " ".join([normalize_header_cell(cell).lower() for cell in row if cell])
        n_matches = sum(1 for kw in keywords if kw in row_text)
        has_amount_kw = "d\u00e9p\u00f4ts" in row_text or "cr\u00e9dit" in row_text
        if n_matches >= 2 and has_amount_kw:
            return index
    return None


def expand_multiline_table(table: list[list[str]]) -> list[list[str]]:
    """Certains rapports BAM (pas de quadrillage de cellules visible dans le
    PDF) sont extraits par pdfplumber comme une poignee de "grosses"
    cellules multi-lignes -- une valeur par ligne physique du tableau,
    empilees avec des "\\n" -- au lieu d'une ligne par enregistrement.
    Verifie en direct sur le rapport "localites" : la ligne de donnees
    entiere d'une page (jusqu'a 50 localites) arrivait comme UNE SEULE ligne
    de table avec 50 valeurs empilees par cellule.

    Scanne TOUTES les lignes (pas seulement un index fixe suppose a
    l'avance : l'ancienne version ne regardait que table[2], ce qui ratait
    ce cas quand l'en-tete ne fait qu'une seule ligne) et explose celles
    dont les cellules non vides ont toutes le MEME nombre de lignes > 1.
    Les lignes deja normales (1 valeur/cellule) ou heterogenes (l'en-tete,
    qui melange souvent 1 et 2 lignes selon les colonnes) sont laissees
    telles quelles.
    """
    expanded: list[list[str]] = []
    for row in table:
        split_columns = [normalize_string(cell).splitlines() if cell else [] for cell in row]
        counts = [len(lines) for lines in split_columns if lines]
        if not counts or max(counts) <= 1 or len(set(counts)) != 1:
            expanded.append(row)
            continue
        n_lines = counts[0]
        for line_index in range(n_lines):
            expanded.append([
                split_columns[col_index][line_index] if line_index < len(split_columns[col_index]) else None
                for col_index in range(len(row))
            ])
    return expanded


# Lignes d'agregation (pas un enregistrement individuel) presentes dans
# certains tableaux BAM -- verifie en direct sur le rapport "localites"
# ("SOUS-TOTAL", "AUTRES LOCALITES (*)", "TOTAL" en fin de tableau).
AGGREGATE_ROW_MARKERS = ("total", "autres localites")


def extract_rows_from_table(table: list[list[str]]) -> list[dict[str, Any]]:
    if not table:
        return []

    table = expand_multiline_table(table)
    header_index = find_header_index(table)
    if header_index is None:
        return []
    n_header_rows = header_row_count(table, header_index)
    header_rows = table[header_index : header_index + n_header_rows]
    headers = merge_header_rows(header_rows)
    if not headers:
        headers = [normalize_header_cell(cell) for cell in table[header_index]]
    headers = [normalize_header_cell(cell) for cell in headers]

    rows: list[dict[str, Any]] = []
    data_rows = table[header_index + n_header_rows :]

    for row in data_rows:
        first_cell = normalize_string(row[0]) if row else ""
        first_cell_norm = remove_accents(first_cell.lower())
        if not first_cell or any(marker in first_cell_norm for marker in AGGREGATE_ROW_MARKERS):
            continue
        values = [normalize_string(cell) for cell in row]
        record: dict[str, Any] = {}
        for index, heading in enumerate(headers):
            if index >= len(values):
                break
            record[heading or f"column_{index}"] = normalize_number(values[index]) if re.search(r"\d", values[index]) else values[index]
        if any(record.values()):
            rows.append(record)

    return rows


def extract_records_from_pdf(pdf_path: Path, report: PdfReportLink, verbose: bool = False) -> list[ScrapedRow]:
    records: list[ScrapedRow] = []
    periode = extract_periode_from_title(report.title)
    with pdfplumber.open(pdf_path) as pdf:
        for page_index, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            if verbose:
                print(f"Page {page_index} found {len(tables)} table(s)")
            for table in tables:
                if not table:
                    continue
                row_dicts = extract_rows_from_table(table)
                for row_number, row_data in enumerate(row_dicts, start=1):
                    if periode and "periode" not in {normalize_field_name(k) for k in row_data}:
                        row_data = {"periode": periode, **row_data}
                    records.append(
                        ScrapedRow(
                            report_title=report.title,
                            report_url=report.url,
                            pdf_filename=pdf_path.name,
                            page_number=page_index,
                            row_number=row_number,
                            data=row_data,
                        )
                    )
    return records


# --------------------------------------------------------------------------
# Extraction depuis un fichier .xlsx (series statistiques monetaires
# credit_objet_economique / credit_secteur_institutionnel). Verifie en
# direct sur les 2 fichiers reels bkam.ma (openpyxl, cellule par cellule) :
# ce N'EST PAS un tableau "en-tete + lignes de donnees" classique. C'est une
# MATRICE LARGE transposee : ligne 1 = titre, ligne 2 = unite ("Encours en
# MDH"), ligne 3 = en-tete de DATES (une colonne par fin de mois depuis
# 2001, colonne A vide), puis une ligne par categorie (compte/agregat en
# colonne A, une valeur par mois dans les colonnes suivantes). Un lecteur
# pandas.read_excel(header=0) classique produit ~294 colonnes "Unnamed: N"
# inexploitables -- d'ou ce parseur dedie qui detecte la ligne d'en-tete
# (recherche de dates, pas position fixe) puis "depivote" en format long
# (categorie, periode, encours_mdh), une ligne par (categorie x mois).
# --------------------------------------------------------------------------

def extract_wide_date_matrix_from_xlsx(xlsx_path: Path, report: PdfReportLink, verbose: bool = False) -> list[ScrapedRow]:
    import openpyxl
    from datetime import datetime as dt

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row_idx = None
    for r in range(1, min(11, ws.max_row + 1)):
        row_cells = [ws.cell(row=r, column=c).value for c in range(2, ws.max_column + 1)]
        n_dates = sum(1 for v in row_cells if isinstance(v, dt))
        if n_dates >= max(3, len(row_cells) // 2):
            header_row_idx = r
            break

    if header_row_idx is None:
        if verbose:
            logger.warning(
                "extract_wide_date_matrix_from_xlsx: aucune ligne d'en-tete "
                "de dates trouvee dans %s (structure inattendue -- fichier "
                "peut-etre change de format cote bkam.ma)",
                xlsx_path.name,
            )
        return []

    date_columns = [
        (c, ws.cell(row=header_row_idx, column=c).value.date().isoformat())
        for c in range(2, ws.max_column + 1)
        if isinstance(ws.cell(row=header_row_idx, column=c).value, dt)
    ]

    records: list[ScrapedRow] = []
    row_number = 0
    for r in range(header_row_idx + 1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        label = normalize_string(label).strip() if label else ""
        if not label:
            continue
        for col_idx, date_str in date_columns:
            value = ws.cell(row=r, column=col_idx).value
            if value is None:
                continue
            row_number += 1
            records.append(
                ScrapedRow(
                    report_title=report.title,
                    report_url=report.url,
                    pdf_filename=xlsx_path.name,
                    page_number=1,
                    row_number=row_number,
                    data={"categorie": label, "periode": date_str, "encours_mdh": value},
                )
            )

    if verbose:
        logger.info(
            "xlsx matrice large '%s' : %d ligne(s) extraite(s) (%d colonne(s) de dates)",
            xlsx_path.name, len(records), len(date_columns),
        )
    return records


# --------------------------------------------------------------------------
# NOUVEAU : extraction "densité bancaire" par regex sur le texte (pas les
# tableaux) du Rapport annuel de supervision bancaire.
# --------------------------------------------------------------------------

DENSITE_BANCAIRE_PATTERNS = [
    # "densité bancaire, mesurée par ... par agence, ressort à 4.709"
    re.compile(
        r"densit[ée]\s+bancaire[^.]{0,120}?ressort\s*à\s*([\d.,\s]+)",
        re.IGNORECASE,
    ),
]

NOMBRE_AGENCES_PATTERNS = [
    # "le nombre d'agences bancaires s'est ... pour ressortir à 5.692"
    re.compile(
        r"nombre\s+d.agences\s+bancaires[^.]{0,120}?ressortir\s*à\s*([\d.,\s]+)",
        re.IGNORECASE,
    ),
]

AGENCES_POUR_10000_HAB_PATTERNS = [
    # La formulation varie d'une edition a l'autre (verifie en direct,
    # edition 2024 : "... est reste stable a 2,1 agences", pas "etabli a")
    # -- accepte plusieurs verbes/tournures usuels de ce type de rapport.
    re.compile(
        r"nombre\s+d.agences\s+pour\s+10\s*\.?\s*000\s+habitants"
        r"[^.]{0,60}?(?:établi|ressort|resté|restée?|maintenu|situé)\w*\s*(?:stable\s*)?à\s*([\d.,]+)",
        re.IGNORECASE,
    ),
]

ANNEE_RAPPORT_PATTERN = re.compile(r"(19|20)\d{2}")


def _first_match(patterns: list[re.Pattern], text: str) -> str | None:
    for pattern in patterns:
        match = pattern.search(text)
        if match:
            return match.group(1).strip()
    return None


def extract_densite_bancaire_from_pdf(pdf_path: Path, report: PdfReportLink, verbose: bool = False) -> list[ScrapedRow]:
    """Parcourt le texte (pas les tableaux) du rapport annuel et en extrait
    les indicateurs de réseau bancaire par regex. Une seule "ligne" de
    résultat par rapport (contrairement aux autres sections qui produisent
    une ligne par ligne de tableau)."""
    full_text_parts: list[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        # Ces indicateurs se trouvent typiquement dans les toutes premières
        # pages du chapitre "Structure du système bancaire" -- on limite la
        # lecture aux ~40 premières pages pour rester rapide, tout en
        # couvrant largement où que tombe ce chapitre.
        for page in pdf.pages[:40]:
            text = page.extract_text() or ""
            full_text_parts.append(text)
    full_text = "\n".join(full_text_parts)

    densite = _first_match(DENSITE_BANCAIRE_PATTERNS, full_text)
    nombre_agences = _first_match(NOMBRE_AGENCES_PATTERNS, full_text)
    agences_10000hab = _first_match(AGENCES_POUR_10000_HAB_PATTERNS, full_text)

    # report.title est generique ("(PDF)") sur cette page -- verifie en
    # direct : le texte des liens ne porte aucune annee. Le nom de fichier,
    # lui, en porte toujours une ("Rapport DSB 2024.pdf", "Rapport SB
    # 2019.pdf", ...) -- on cherche donc l'annee dans le nom de fichier
    # d'abord, avec le titre comme repli.
    annee_match = ANNEE_RAPPORT_PATTERN.search(pdf_path.name) or ANNEE_RAPPORT_PATTERN.search(report.title)
    annee = annee_match.group(0) if annee_match else None

    if not any([densite, nombre_agences, agences_10000hab]):
        if verbose:
            logger.warning(
                "Aucun indicateur de densité bancaire trouvé dans '%s' -- "
                "les regex DENSITE_BANCAIRE_PATTERNS/NOMBRE_AGENCES_PATTERNS "
                "ont probablement besoin d'être ajustées à la formulation "
                "exacte de cette édition du rapport.",
                report.title,
            )
        return []

    return [
        ScrapedRow(
            report_title=report.title,
            report_url=report.url,
            pdf_filename=pdf_path.name,
            page_number=0,
            row_number=1,
            data={
                "annee_rapport": annee,
                "nombre_agences_bancaires": normalize_number(nombre_agences) if nombre_agences else None,
                "densite_bancaire": normalize_number(densite) if densite else None,
                "agences_pour_10000_habitants": normalize_number(agences_10000hab) if agences_10000hab else None,
            },
        )
    ]


def save_records(records: list[ScrapedRow], output_path: Path, output_format: str, normalize: bool = False, sqlite_table: str | None = None) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    normalized_records = records

    if normalize:
        for record in normalized_records:
            record.data = normalize_data_for_db(record.data)

    if output_format == "json":
        with output_path.open("w", encoding="utf-8") as handle:
            json.dump([asdict(record) for record in normalized_records], handle, indent=2, ensure_ascii=False)
        return

    if output_format == "sqlite":
        save_records_to_sqlite(normalized_records, output_path, sqlite_table or "bank_almaghreb")
        return

    headers = build_csv_headers(normalized_records)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for record in normalized_records:
            row = {
                "report_title": record.report_title,
                "report_url": record.report_url,
                "pdf_filename": record.pdf_filename,
                "page_number": record.page_number,
                "row_number": record.row_number,
            }
            row.update(record.data)
            writer.writerow(row)


def sanitize_identifier(name: str) -> str:
    """Make a safe SQLite identifier from user input (used unescaped in DDL/DML)."""
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", name).strip("_")
    if not cleaned:
        cleaned = "bank_almaghreb"
    if cleaned[0].isdigit():
        cleaned = f"t_{cleaned}"
    return cleaned


def infer_sql_type(values: list[Any]) -> str:
    has_real = False
    for value in values:
        if value is None:
            continue
        if isinstance(value, int):
            continue
        if isinstance(value, float):
            has_real = True
            continue
        return "TEXT"
    return "REAL" if has_real else "INTEGER"


def save_records_to_sqlite(records: list[ScrapedRow], sqlite_path: Path, table_name: str) -> None:
    if not records:
        return

    table_name = sanitize_identifier(table_name)
    rows = []
    columns: set[str] = set()
    for record in records:
        record_data = normalize_data_for_db(record.data)
        row = {
            "report_title": record.report_title,
            "report_url": record.report_url,
            "pdf_filename": record.pdf_filename,
            "page_number": record.page_number,
            "row_number": record.row_number,
            **record_data,
        }
        columns.update(row.keys())
        rows.append(row)

    sqlite_path.unlink(missing_ok=True)
    with sqlite3.connect(sqlite_path) as conn:
        cursor = conn.cursor()
        column_types: dict[str, str] = {}
        for column in sorted(columns):
            values = [row.get(column) for row in rows]
            if column in {"report_title", "report_url", "pdf_filename"}:
                column_types[column] = "TEXT"
            elif column in {"page_number", "row_number"}:
                column_types[column] = "INTEGER"
            else:
                column_types[column] = infer_sql_type(values)

        columns_sql = ", ".join(f"{column} {column_types[column]}" for column in sorted(columns))
        cursor.execute(f"CREATE TABLE {table_name} ({columns_sql})")

        placeholders = ", ".join("?" for _ in sorted(columns))
        insert_sql = f"INSERT INTO {table_name} ({', '.join(sorted(columns))}) VALUES ({placeholders})"
        for row in rows:
            values = [row.get(column) for column in sorted(columns)]
            cursor.execute(insert_sql, values)
        conn.commit()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape Bank Al Maghrib statistics PDFs and extract table rows.")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--section",
        choices=SECTION_PAGES,
        help="Select a single statistics section to scrape.",
    )
    group.add_argument(
        "--all",
        action="store_true",
        help="Scrape every section in ALL_SECTIONS, one CSV per section (used by pipeline.py).",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output CSV/JSON/SQLite path (--section mode only; --all computes one path per section).",
    )
    parser.add_argument(
        "--format",
        choices=["csv", "json", "sqlite"],
        default="csv",
        help="Output data format.",
    )
    parser.add_argument(
        "--normalize",
        action="store_true",
        help="Normalize extracted table headers and values for database import "
             "(always on in --all mode: the bronze SQL layer expects snake_case headers).",
    )
    parser.add_argument(
        "--sqlite-table",
        default="bank_almaghreb",
        help="Table name when writing sqlite output.",
    )
    parser.add_argument(
        "--download-dir",
        default="../../../datasets/bkm/raw/pdf_cache",
        help="Directory to store downloaded PDFs (cache, keyed by filename).",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Limit the number of reports to download and parse (per section).",
    )
    parser.add_argument(
        "--all-reports",
        action="store_true",
        help="Scrape all PDF reports from the selected section, even if they are not filtered by keyword.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="List the PDF reports that would be downloaded/parsed, without downloading or writing output.",
    )
    parser.add_argument("--verbose", action="store_true", help="Print progress messages.")
    args = parser.parse_args()
    if args.section and args.output is None:
        args.output = "../../../datasets/bkm/bank_almaghreb_data.csv"
    return args


def scrape_section(
    section_key: str,
    output_path: Path,
    download_dir: Path,
    session: Session,
    output_format: str,
    normalize: bool,
    sqlite_table: str,
    limit: int | None,
    all_reports: bool,
    dry_run: bool,
    verbose: bool,
) -> int:
    """Scrape one SECTION_PAGES entry end-to-end (discover -> download -> extract -> save).
    Shared by both --section (single) and --all (looped) modes."""
    section = SECTION_PAGES[section_key]
    kind = section.get("kind", "pdf_table")
    accepted_extensions = ACCEPTED_EXTENSIONS_BY_KIND.get(kind, (".pdf",))

    try:
        pdf_links = fetch_pdf_links(
            section["url"],
            session,
            None if all_reports else section.get("filter_keywords"),
            accepted_extensions=accepted_extensions,
            exclude_keywords=None if all_reports else section.get("exclude_keywords"),
        )
    except RequestException as exc:
        logger.error("Could not reach %s: %s", section["url"], exc)
        return 1

    if not pdf_links:
        print(f"No PDF/XLSX links found for section {section_key}.")
        return 1

    if limit:
        pdf_links = pdf_links[:limit]

    if verbose:
        logger.info("Found %d report(s) from %s", len(pdf_links), section["label"])

    if dry_run:
        for report in pdf_links:
            print(f"{report.title}\t{report.url}")
        print(f"({len(pdf_links)} report(s) would be processed)")
        return 0

    scraped_records: list[ScrapedRow] = []
    for report in pdf_links:
        try:
            file_path = download_pdf(report, download_dir, session, verbose=verbose)
            if kind == "pdf_text_regex":
                scraped_records.extend(extract_densite_bancaire_from_pdf(file_path, report, verbose=verbose))
            elif file_path.suffix.lower() in (".xlsx", ".xls"):
                scraped_records.extend(extract_wide_date_matrix_from_xlsx(file_path, report, verbose=verbose))
            else:
                scraped_records.extend(extract_records_from_pdf(file_path, report, verbose=verbose))
        except (RequestException, ValueError) as exc:
            logger.error("Error processing %s: %s", report.url, exc)
        except Exception as exc:  # unexpected parsing errors: keep going, but surface clearly
            logger.exception("Unexpected error processing %s: %s", report.url, exc)

    if not scraped_records:
        print(f"[WARN] {section_key}: 0 row extracted, {output_path} not written.")
        return 1

    save_records(
        scraped_records,
        output_path,
        output_format,
        normalize=normalize or output_format == "sqlite",
        sqlite_table=sqlite_table,
    )
    print(f"[OK] {section_key}: wrote {len(scraped_records)} row(s) to {output_path}")
    return 0


def main() -> int:
    args = parse_args()
    logging.basicConfig(
        level=logging.INFO if args.verbose else logging.WARNING,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    session = requests.Session()
    session.headers.update(DEFAULT_HEADERS)
    download_dir = Path(args.download_dir)

    if args.section:
        return scrape_section(
            args.section,
            Path(args.output),
            download_dir,
            session,
            args.format,
            args.normalize,
            args.sqlite_table,
            args.limit,
            args.all_reports,
            args.dry_run,
            args.verbose,
        )

    # --all : boucle sur ALL_SECTIONS, un CSV par section, toujours normalise.
    # Limite par defaut par section (si --limit n'est pas fourni explicitement) :
    # les sections mensuelles (regional_credit/credits_depots_localites)
    # n'ont besoin que d'un historique recent au premier run (le chargement
    # bronze est incremental, un futur --scrape en ajoutera plus) ;
    # densite_bancaire est un indicateur ANNUEL, 1 seul rapport (le plus
    # recent) suffit -- inutile de re-parser 20 PDF de ~150 pages a chaque run.
    default_limits = {
        "regional_credit": 12,
        "credits_depots_localites": 12,
        "credit_objet_economique": 1,
        "credit_secteur_institutionnel": 1,
        "densite_bancaire": 1,
    }

    exit_code = 0
    for section_key in ALL_SECTIONS:
        output_path = Path("../../../datasets/bkm") / SECTION_OUTPUT_FILENAMES[section_key]
        code = scrape_section(
            section_key,
            output_path,
            download_dir,
            session,
            "csv",
            True,
            args.sqlite_table,
            args.limit or default_limits[section_key],
            args.all_reports,
            args.dry_run,
            args.verbose,
        )
        if code != 0:
            exit_code = code
        if section_key != ALL_SECTIONS[-1]:
            time.sleep(2)

    return exit_code


if __name__ == "__main__":
    raise SystemExit(main())